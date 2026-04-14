"""
Financial Model Generator – Final
Fixed in this version:
  - CRITICAL: Cash formula off-by-one (was Y(N-1) cash flow, now YN) → balance sheet now balances
  - Sources & Uses formula: added missing $ anchors
  - Break-even year KPI: fixed and confirmed present
  - Dead code removed (units_note)
  - Dashboard helper row hidden
  - apply_money helper cleaned up
  - Insights commentary escaping fixed
  - Row heights set on all sheets for clean visual spacing
  - All-around UX polish pass
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ─────────────────────────────────────────────
# 1. Config
# ─────────────────────────────────────────────
cfg_path = Path(__file__).parent / "config.json"
with open(cfg_path) as f:
    cfg = json.load(f)

for key in ["forecast_years", "base_revenue", "initial_cash", "initial_ppe", "initial_debt"]:
    if key not in cfg:
        raise ValueError(f"Missing required config key: {key}")

YEARS        = cfg["forecast_years"]
BASE_REV     = cfg["base_revenue"]
SCENARIOS    = cfg["scenarios"]
ASSUMPTIONS  = cfg["assumptions"]
INITIAL_CASH = cfg["initial_cash"]
INITIAL_PPE  = cfg["initial_ppe"]
INITIAL_DEBT = cfg["initial_debt"]

RD_PCT          = ASSUMPTIONS["rd_pct"]
SGA_PCT         = ASSUMPTIONS["sga_pct"]
TAX_RATE        = ASSUMPTIONS["tax_rate"]
DEPR_RATE       = ASSUMPTIONS["depreciation_rate"]
CAPEX_PCT       = ASSUMPTIONS["capex_pct"]
DSO             = ASSUMPTIONS["dso_days"]
DIO             = ASSUMPTIONS["dio_days"]
DPO             = ASSUMPTIONS["dpo_days"]
INTEREST_RATE   = ASSUMPTIONS["interest_rate"]
PRINCIPAL_REPAY = ASSUMPTIONS["principal_repayment"]
WACC_VAL        = ASSUMPTIONS["wacc"]
PERP_GROWTH     = ASSUMPTIONS["perpetual_growth"]

# ─────────────────────────────────────────────
# 2. Workbook & style helpers
# ─────────────────────────────────────────────
wb = Workbook()

def CL(col): return get_column_letter(col)

DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
YELLOW     = "FFFF00"
WHITE      = "FFFFFF"

_thin = Side(style="thin")
_med  = Side(style="medium")
border_box    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
border_bottom = Border(bottom=_med)
border_top    = Border(top=_med)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Arial", italic=italic)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def apply_header(cell):
    cell.font      = _font(bold=True, color=WHITE)
    cell.fill      = _fill(DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border_box

def apply_subheader(cell):
    cell.font      = _font(bold=True, color=WHITE)
    cell.fill      = _fill(MID_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = border_box

def apply_input(cell):
    """Blue text on yellow — hardcoded input (industry standard)."""
    cell.font       = _font(color="0000FF")
    cell.fill       = _fill(YELLOW)
    cell.protection = Protection(locked=False)

def apply_formula(cell):
    """Black text — calculated cell."""
    cell.font = _font()

def apply_link(cell):
    """Green text — cross-sheet link."""
    cell.font = _font(color="008000")

def apply_flag(cell):
    """Red italic — warning flag."""
    cell.font = _font(color="9C0006", italic=True, size=9)

def apply_money(cell, is_input=False):
    cell.number_format = '#,##0_);(#,##0);"-"'
    apply_input(cell) if is_input else apply_formula(cell)

def apply_percent(cell, is_input=False):
    cell.number_format = "0.0%"
    apply_input(cell) if is_input else apply_formula(cell)

def apply_label(cell, bold=False, indent=1, color="000000"):
    cell.font      = _font(bold=bold, color=color)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)

def ref(sheet, row, col):
    """Absolute cross-sheet reference: 'Sheet'!$C$5"""
    return f"'{sheet}'!${CL(col)}${row}"

def set_col_widths(ws, label_col_width=30, data_col_width=14, n_data_cols=None):
    ws.column_dimensions["A"].width = label_col_width
    end = (n_data_cols or YEARS) + 1
    for c in range(2, end + 1):
        ws.column_dimensions[CL(c)].width = data_col_width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

# ─────────────────────────────────────────────
# 3. Row maps
# ─────────────────────────────────────────────
pl_rows = {
    "Revenue":      2,
    "COGS":         3,
    "Gross Profit": 4,
    "R&D":          5,
    "SG&A":         6,
    "EBITDA":       7,
    "Depreciation": 8,
    "EBIT":         9,
    "Interest":     10,
    "EBT":          11,
    "Taxes":        12,
    "Net Income":   13,
}

bs_rows = {
    "Cash":                      2,
    "Accounts Receivable":       3,
    "Inventory":                 4,
    "PP&E":                      5,
    "Total Assets":              6,
    # row 7: separator
    "Accounts Payable":          8,
    "Debt":                      9,
    "Equity Plug":               10,
    "Retained Earnings":         11,
    "Total Liabilities & Equity":12,
}
BS_CHECK_ROW = 14

cf_rows = {
    "Net Income":           2,
    "Depreciation":         3,
    "Change in AR":         4,
    "Change in Inventory":  5,
    "Change in AP":         6,
    "Cash from Operations": 7,
    "Capex":                8,
    "Cash from Investing":  9,
    "Debt Repayment":       10,
    "Cash from Financing":  11,
    "Net Change in Cash":   12,
}

BS = "Balance Sheet"
CF = "Cash Flow"
CP = "Control Panel"

# ─────────────────────────────────────────────
# 4. Control Panel
# ─────────────────────────────────────────────
cp = wb.create_sheet(CP)

CP_ASSUMP_START = 3

# Assumption order — wacc must come BEFORE perpetual_growth so the flag formula works
assumption_order = [
    ("rd_pct",              "R&D % of Revenue",          RD_PCT,          "percent"),
    ("sga_pct",             "SG&A % of Revenue",         SGA_PCT,         "percent"),
    ("tax_rate",            "Tax Rate",                   TAX_RATE,        "percent"),
    ("depreciation_rate",   "Depreciation Rate (PP&E)",  DEPR_RATE,       "percent"),
    ("capex_pct",           "Capex % of Revenue",        CAPEX_PCT,       "percent"),
    ("dso_days",            "DSO (days)",                DSO,             "number"),
    ("dio_days",            "DIO (days)",                DIO,             "number"),
    ("dpo_days",            "DPO (days)",                DPO,             "number"),
    ("interest_rate",       "Interest Rate on Debt",     INTEREST_RATE,   "percent"),
    ("principal_repayment", "Annual Debt Repayment",     PRINCIPAL_REPAY, "number"),
    ("wacc",                "WACC",                      WACC_VAL,        "percent"),
    ("perpetual_growth",    "Perpetual Growth Rate",     PERP_GROWTH,     "percent"),
]
CP_ASSUMP_ROW = {}

# Title
cp["A1"] = "FINANCIAL MODEL – CONTROL PANEL"
cp["A1"].font = _font(bold=True, size=14, color=DARK_BLUE)
cp.merge_cells("A1:F1")
cp["A1"].alignment = Alignment(horizontal="center")
set_row_height(cp, 1, 28)

# Column headers
for c, h in enumerate(["Assumption", "Value", "Flag"], start=1):
    apply_header(cp.cell(row=2, column=c, value=h))

# Assumption rows
for i, (key, label, val, fmt) in enumerate(assumption_order):
    row = CP_ASSUMP_START + i
    CP_ASSUMP_ROW[key] = row
    cp.cell(row=row, column=1, value=label); apply_label(cp.cell(row=row, column=1))
    cell_b = cp.cell(row=row, column=2, value=val)
    apply_percent(cell_b, is_input=True) if fmt == "percent" else apply_money(cell_b, is_input=True)

# Assumption flags (column C)
FLAG_DEF = {
    "rd_pct":            (">0.20",  "⚠ R&D > 20% is very high"),
    "sga_pct":           (">0.30",  "⚠ SG&A > 30% — watch opex"),
    "tax_rate":          (">0.35",  "⚠ Tax rate above 35%"),
    "depreciation_rate": (">0.20",  "⚠ Very rapid depreciation"),
    "capex_pct":         (">0.15",  "⚠ Capex > 15% of revenue"),
    "interest_rate":     (">0.10",  "⚠ Interest rate above 10%"),
    "wacc":              (">0.15",  "⚠ WACC > 15% is very high"),
}
for key, (cond, msg) in FLAG_DEF.items():
    if key not in CP_ASSUMP_ROW:
        continue
    row = CP_ASSUMP_ROW[key]
    cp.cell(row=row, column=3, value=f'=IF($B${row}{cond},"{msg}","")')
    apply_flag(cp.cell(row=row, column=3))

# perpetual_growth flag: g >= WACC makes TV infinite
pg_row   = CP_ASSUMP_ROW["perpetual_growth"]
wacc_row = CP_ASSUMP_ROW["wacc"]
cp.cell(row=pg_row, column=3,
        value=f'=IF($B${pg_row}>=$B${wacc_row},"⚠ g ≥ WACC → Terminal Value invalid","")')
apply_flag(cp.cell(row=pg_row, column=3))

# ── Scenario table ──
SCEN_HDR_ROW    = CP_ASSUMP_START + len(assumption_order) + 2
SCEN_DATA_START = SCEN_HDR_ROW + 1
scen_names      = list(SCENARIOS.keys())

for c, lbl in enumerate(["Scenario", "Revenue Growth", "COGS %"], start=1):
    apply_header(cp.cell(row=SCEN_HDR_ROW, column=c, value=lbl))

for i, (sname, sdata) in enumerate(SCENARIOS.items()):
    r = SCEN_DATA_START + i
    cp.cell(row=r, column=1, value=sname);                    apply_label(cp.cell(row=r, column=1))
    cp.cell(row=r, column=2, value=sdata["revenue_growth"]);  apply_percent(cp.cell(row=r, column=2), is_input=True)
    cp.cell(row=r, column=3, value=sdata["cogs_pct"]);        apply_percent(cp.cell(row=r, column=3), is_input=True)
    # Active indicator
    cp.cell(row=r, column=4, value=f'=IF($B${SCEN_DATA_START+len(SCENARIOS)+2}=$A${r},"◀ active","")')
    apply_flag(cp.cell(row=r, column=4))
    cp.cell(row=r, column=4).font = _font(color="008000", italic=True, size=9)
    for c in range(1, 4):
        cp.cell(row=r, column=c).protection = Protection(locked=False)

# ── Scenario selector ──
SEL_ROW = SCEN_DATA_START + len(SCENARIOS) + 2
cp.cell(row=SEL_ROW, column=1, value="▶  Active Scenario")
apply_subheader(cp.cell(row=SEL_ROW, column=1))
cp.cell(row=SEL_ROW, column=2, value="Base")
apply_input(cp.cell(row=SEL_ROW, column=2))

# Range-based dropdown (scalable: adding scenario rows auto-extends it)
dv = DataValidation(
    type="list",
    formula1=f"=$A${SCEN_DATA_START}:$A${SCEN_DATA_START + len(SCENARIOS) - 1}",
    showDropDown=False
)
cp.add_data_validation(dv)
dv.add(f"B{SEL_ROW}")
CP_SEL_ADDR = f"$B${SEL_ROW}"

# Effective growth & COGS
EFF_GROWTH_ROW = SEL_ROW + 1
EFF_COGS_ROW   = SEL_ROW + 2
sr_names  = f"$A${SCEN_DATA_START}:$A${SCEN_DATA_START + len(SCENARIOS) - 1}"
sr_growth = f"$B${SCEN_DATA_START}:$B${SCEN_DATA_START + len(SCENARIOS) - 1}"
sr_cogs   = f"$C${SCEN_DATA_START}:$C${SCEN_DATA_START + len(SCENARIOS) - 1}"

cp.cell(row=EFF_GROWTH_ROW, column=1, value="Effective Revenue Growth")
cp.cell(row=EFF_GROWTH_ROW, column=2,
        value=f"=INDEX({sr_growth},MATCH({CP_SEL_ADDR},{sr_names},0))")
apply_label(cp.cell(row=EFF_GROWTH_ROW, column=1), bold=True)
apply_percent(cp.cell(row=EFF_GROWTH_ROW, column=2))

cp.cell(row=EFF_COGS_ROW, column=1, value="Effective COGS %")
cp.cell(row=EFF_COGS_ROW, column=2,
        value=f"=INDEX({sr_cogs},MATCH({CP_SEL_ADDR},{sr_names},0))")
apply_label(cp.cell(row=EFF_COGS_ROW, column=1), bold=True)
apply_percent(cp.cell(row=EFF_COGS_ROW, column=2))

EFF_GROWTH_REF = f"'{CP}'!$B${EFF_GROWTH_ROW}"
EFF_COGS_REF   = f"'{CP}'!$B${EFF_COGS_ROW}"

def cp_ref(key):
    return f"'{CP}'!$B${CP_ASSUMP_ROW[key]}"

# Model health (filled after BS is built, BC_RANGE defined there)
HEALTH_ROW = EFF_COGS_ROW + 2

# Column widths
cp.column_dimensions["A"].width = 30
cp.column_dimensions["B"].width = 18
cp.column_dimensions["C"].width = 36
cp.column_dimensions["D"].width = 12

# ─────────────────────────────────────────────
# 5. P&L
# ─────────────────────────────────────────────
pl = wb.create_sheet("P&L")
for c, h in enumerate(["Line Item (USD)"] + [f"Year {y}" for y in range(1, YEARS+1)], 1):
    apply_header(pl.cell(row=1, column=c, value=h))

pl_label_map = {
    "Revenue":      ("Revenue",                       False, False),
    "COGS":         ("  Cost of Goods Sold",          False, False),
    "Gross Profit": ("Gross Profit",                  True,  True),
    "R&D":          ("  R&D Expense",                 False, False),
    "SG&A":         ("  SG&A Expense",                False, False),
    "EBITDA":       ("EBITDA",                        True,  True),
    "Depreciation": ("  Depreciation & Amortization", False, False),
    "EBIT":         ("EBIT",                          True,  True),
    "Interest":     ("  Interest Expense",            False, False),
    "EBT":          ("EBT",                           True,  True),
    "Taxes":        ("  Income Taxes",                False, False),
    "Net Income":   ("Net Income",                    True,  True),
}
for name, (label, bold, border) in pl_label_map.items():
    cell = pl.cell(row=pl_rows[name], column=1, value=label)
    apply_label(cell, bold=bold)
    if border:
        cell.border = border_bottom

def pl_ref(row_name, col):
    return ref("P&L", pl_rows[row_name], col)

# Year 1 Revenue — input
apply_money(pl.cell(row=pl_rows["Revenue"], column=2, value=BASE_REV), is_input=True)

# Year 2..N Revenue
for col in range(3, YEARS+2):
    pl.cell(row=pl_rows["Revenue"], column=col,
            value=f"={pl_ref('Revenue', col-1)}*(1+{EFF_GROWTH_REF})")
    apply_money(pl.cell(row=pl_rows["Revenue"], column=col))

for col in range(2, YEARS+2):
    rev = pl_ref("Revenue", col)

    pl.cell(row=pl_rows["COGS"], column=col,     value=f"={rev}*{EFF_COGS_REF}")
    apply_money(pl.cell(row=pl_rows["COGS"], column=col))

    pl.cell(row=pl_rows["Gross Profit"], column=col,
            value=f"={pl_ref('Revenue',col)}-{pl_ref('COGS',col)}")
    apply_money(pl.cell(row=pl_rows["Gross Profit"], column=col))

    pl.cell(row=pl_rows["R&D"],  column=col, value=f"={rev}*{cp_ref('rd_pct')}")
    apply_money(pl.cell(row=pl_rows["R&D"], column=col))

    pl.cell(row=pl_rows["SG&A"], column=col, value=f"={rev}*{cp_ref('sga_pct')}")
    apply_money(pl.cell(row=pl_rows["SG&A"], column=col))

    pl.cell(row=pl_rows["EBITDA"], column=col,
            value=f"={pl_ref('Gross Profit',col)}-{pl_ref('R&D',col)}-{pl_ref('SG&A',col)}")
    apply_money(pl.cell(row=pl_rows["EBITDA"], column=col))

    # Depreciation: opening PP&E * rate (no circular ref: uses prior-year BS)
    opening_ppe = INITIAL_PPE if col == 2 else ref(BS, bs_rows["PP&E"], col-1)
    pl.cell(row=pl_rows["Depreciation"], column=col,
            value=f"={opening_ppe}*{cp_ref('depreciation_rate')}")
    apply_money(pl.cell(row=pl_rows["Depreciation"], column=col))

    pl.cell(row=pl_rows["EBIT"], column=col,
            value=f"={pl_ref('EBITDA',col)}-{pl_ref('Depreciation',col)}")
    apply_money(pl.cell(row=pl_rows["EBIT"], column=col))

    # Interest: opening debt only — safe, no circular reference
    opening_debt = INITIAL_DEBT if col == 2 else ref(BS, bs_rows["Debt"], col-1)
    pl.cell(row=pl_rows["Interest"], column=col,
            value=f"={opening_debt}*{cp_ref('interest_rate')}")
    apply_money(pl.cell(row=pl_rows["Interest"], column=col))

    pl.cell(row=pl_rows["EBT"], column=col,
            value=f"={pl_ref('EBIT',col)}-{pl_ref('Interest',col)}")
    apply_money(pl.cell(row=pl_rows["EBT"], column=col))

    # Taxes: MAX(EBT × rate, 0) — no negative taxes on losses
    pl.cell(row=pl_rows["Taxes"], column=col,
            value=f"=MAX({pl_ref('EBT',col)}*{cp_ref('tax_rate')},0)")
    apply_money(pl.cell(row=pl_rows["Taxes"], column=col))

    pl.cell(row=pl_rows["Net Income"], column=col,
            value=f"={pl_ref('EBT',col)}-{pl_ref('Taxes',col)}")
    apply_money(pl.cell(row=pl_rows["Net Income"], column=col))

# ─────────────────────────────────────────────
# 6. Balance Sheet
# ─────────────────────────────────────────────
bs_ws = wb.create_sheet("Balance Sheet")
for c, h in enumerate(["Line Item (USD)"] + [f"Year {y}" for y in range(1, YEARS+1)], 1):
    apply_header(bs_ws.cell(row=1, column=c, value=h))

bs_label_map = {
    "Cash":                       ("Cash & Equivalents",         False),
    "Accounts Receivable":        ("  Accounts Receivable",      False),
    "Inventory":                  ("  Inventory",                False),
    "PP&E":                       ("  PP&E, net",                False),
    "Total Assets":               ("Total Assets",               True),
    "Accounts Payable":           ("  Accounts Payable",         False),
    "Debt":                       ("  Debt",                     False),
    "Equity Plug":                ("  Paid-In Capital",          False),
    "Retained Earnings":          ("  Retained Earnings",        False),
    "Total Liabilities & Equity": ("Total Liabilities & Equity", True),
}
for name, (label, bold) in bs_label_map.items():
    cell = bs_ws.cell(row=bs_rows[name], column=1, value=label)
    apply_label(cell, bold=bold)
    if bold:
        cell.border = border_bottom

bs_ws.cell(row=7, column=1, value="LIABILITIES & EQUITY")
apply_label(bs_ws.cell(row=7, column=1), bold=True, color="595959")
bs_ws.cell(row=7, column=1).font = _font(bold=True, color="595959", size=9)

# ── Cash  FIX: Cash YN = Cash Y(N-1) + NChg YN (SAME column) ──
apply_money(bs_ws.cell(row=bs_rows["Cash"], column=2, value=INITIAL_CASH), is_input=True)
for col in range(3, YEARS+2):
    prev = ref(BS, bs_rows["Cash"], col-1)
    # FIXED: uses CF col (current year), not col-1 (prior year)
    nchg = ref(CF, cf_rows["Net Change in Cash"], col)
    bs_ws.cell(row=bs_rows["Cash"], column=col, value=f"={prev}+{nchg}")
    apply_money(bs_ws.cell(row=bs_rows["Cash"], column=col))

# AR: average-period revenue × DSO/365
for col in range(2, YEARS+2):
    rev_c = pl_ref("Revenue", col)
    rev_p = pl_ref("Revenue", col-1) if col > 2 else pl_ref("Revenue", col)
    bs_ws.cell(row=bs_rows["Accounts Receivable"], column=col,
               value=f"=(({rev_c}+{rev_p})/2)*{cp_ref('dso_days')}/365")
    apply_money(bs_ws.cell(row=bs_rows["Accounts Receivable"], column=col))

# Inventory: average-period COGS × DIO/365
for col in range(2, YEARS+2):
    cogs_c = pl_ref("COGS", col)
    cogs_p = pl_ref("COGS", col-1) if col > 2 else pl_ref("COGS", col)
    bs_ws.cell(row=bs_rows["Inventory"], column=col,
               value=f"=(({cogs_c}+{cogs_p})/2)*{cp_ref('dio_days')}/365")
    apply_money(bs_ws.cell(row=bs_rows["Inventory"], column=col))

# PP&E: Y1=initial, YN = prev - CF_Capex(neg) - Depr
apply_money(bs_ws.cell(row=bs_rows["PP&E"], column=2, value=INITIAL_PPE), is_input=True)
for col in range(3, YEARS+2):
    prev_ppe = ref(BS, bs_rows["PP&E"], col-1)
    cf_capex = ref(CF, cf_rows["Capex"], col)   # negative value
    depr     = pl_ref("Depreciation", col)
    bs_ws.cell(row=bs_rows["PP&E"], column=col,
               value=f"={prev_ppe}-({cf_capex})-{depr}")
    apply_money(bs_ws.cell(row=bs_rows["PP&E"], column=col))

# Total Assets
for col in range(2, YEARS+2):
    parts = "+".join(ref(BS, bs_rows[k], col)
                     for k in ("Cash", "Accounts Receivable", "Inventory", "PP&E"))
    bs_ws.cell(row=bs_rows["Total Assets"], column=col, value=f"={parts}")
    apply_money(bs_ws.cell(row=bs_rows["Total Assets"], column=col))

# AP: average-period COGS × DPO/365
for col in range(2, YEARS+2):
    cogs_c = pl_ref("COGS", col)
    cogs_p = pl_ref("COGS", col-1) if col > 2 else pl_ref("COGS", col)
    bs_ws.cell(row=bs_rows["Accounts Payable"], column=col,
               value=f"=(({cogs_c}+{cogs_p})/2)*{cp_ref('dpo_days')}/365")
    apply_money(bs_ws.cell(row=bs_rows["Accounts Payable"], column=col))

# Debt: MAX(0, prev − repayment) — prevents negative debt
apply_money(bs_ws.cell(row=bs_rows["Debt"], column=2, value=INITIAL_DEBT), is_input=True)
for col in range(3, YEARS+2):
    prev_debt = ref(BS, bs_rows["Debt"], col-1)
    bs_ws.cell(row=bs_rows["Debt"], column=col,
               value=f"=MAX(0,{prev_debt}-{cp_ref('principal_repayment')})")
    apply_money(bs_ws.cell(row=bs_rows["Debt"], column=col))

# Retained Earnings: Y1 = NI Y1, YN = prev + NI YN
# (Y1 must capture Year 1 earnings so the plug is calibrated correctly)
for col in range(2, YEARS+2):
    if col == 2:
        bs_ws.cell(row=bs_rows["Retained Earnings"], column=col,
                   value=f"={pl_ref('Net Income', col)}")
    else:
        prev_re = ref(BS, bs_rows["Retained Earnings"], col-1)
        ni_curr = pl_ref("Net Income", col)
        bs_ws.cell(row=bs_rows["Retained Earnings"], column=col,
                   value=f"={prev_re}+{ni_curr}")
    apply_money(bs_ws.cell(row=bs_rows["Retained Earnings"], column=col))

# Equity Plug: Year 1 = TA − AP − Debt − RE (absorbs opening capital)
#              Year 2+ = constant (paid-in capital doesn't change)
col = 2
ta  = ref(BS, bs_rows["Total Assets"], col)
ap  = ref(BS, bs_rows["Accounts Payable"], col)
dbt = ref(BS, bs_rows["Debt"], col)
re  = ref(BS, bs_rows["Retained Earnings"], col)
bs_ws.cell(row=bs_rows["Equity Plug"], column=col, value=f"={ta}-{ap}-{dbt}-{re}")
apply_money(bs_ws.cell(row=bs_rows["Equity Plug"], column=col))
plug_y1 = ref(BS, bs_rows["Equity Plug"], 2)
for col in range(3, YEARS+2):
    bs_ws.cell(row=bs_rows["Equity Plug"], column=col, value=f"={plug_y1}")
    apply_money(bs_ws.cell(row=bs_rows["Equity Plug"], column=col))

# Total L&E
for col in range(2, YEARS+2):
    parts = "+".join(ref(BS, bs_rows[k], col)
                     for k in ("Accounts Payable", "Debt", "Equity Plug", "Retained Earnings"))
    bs_ws.cell(row=bs_rows["Total Liabilities & Equity"], column=col, value=f"={parts}")
    apply_money(bs_ws.cell(row=bs_rows["Total Liabilities & Equity"], column=col))

# Balance Check row
bc_label = bs_ws.cell(row=BS_CHECK_ROW, column=1, value="Balance Check  (must = 0)")
bc_label.font = _font(bold=True, size=9, italic=True)
BC_RANGE = f"$B${BS_CHECK_ROW}:${CL(YEARS+1)}${BS_CHECK_ROW}"

red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
red_font = _font(bold=True, color="9C0006")
grn_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
grn_font = _font(bold=True, color="276221")

for col in range(2, YEARS+2):
    ta = ref(BS, bs_rows["Total Assets"], col)
    le = ref(BS, bs_rows["Total Liabilities & Equity"], col)
    cell = bs_ws.cell(row=BS_CHECK_ROW, column=col, value=f"={ta}-{le}")
    apply_money(cell)
    addr = f"{CL(col)}{BS_CHECK_ROW}"
    bs_ws.conditional_formatting.add(addr,
        CellIsRule(operator="notEqual", formula=["0"], fill=red_fill, font=red_font))
    bs_ws.conditional_formatting.add(addr,
        CellIsRule(operator="equal",    formula=["0"], fill=grn_fill, font=grn_font))

# ─────────────────────────────────────────────
# 7. Cash Flow
# ─────────────────────────────────────────────
cf_ws = wb.create_sheet("Cash Flow")
for c, h in enumerate(["Line Item (USD)"] + [f"Year {y}" for y in range(1, YEARS+1)], 1):
    apply_header(cf_ws.cell(row=1, column=c, value=h))

cf_label_map = {
    "Net Income":           ("Net Income",                     False),
    "Depreciation":         ("  + Depreciation & Amortization",False),
    "Change in AR":         ("  Δ Accounts Receivable",        False),
    "Change in Inventory":  ("  Δ Inventory",                  False),
    "Change in AP":         ("  Δ Accounts Payable",           False),
    "Cash from Operations": ("Cash from Operations",           True),
    "Capex":                ("  Capital Expenditures",         False),
    "Cash from Investing":  ("Cash from Investing",            True),
    "Debt Repayment":       ("  Debt Repayment",               False),
    "Cash from Financing":  ("Cash from Financing",            True),
    "Net Change in Cash":   ("Net Change in Cash",             True),
}
for name, (label, bold) in cf_label_map.items():
    cell = cf_ws.cell(row=cf_rows[name], column=1, value=label)
    apply_label(cell, bold=bold)
    if bold:
        cell.border = border_bottom

for col in range(2, YEARS+2):
    # Net Income — green link from P&L
    cell = cf_ws.cell(row=cf_rows["Net Income"], column=col,
                      value=f"={pl_ref('Net Income', col)}")
    cell.number_format = '#,##0_);(#,##0);"-"'; apply_link(cell)

    # Depreciation — green link from P&L
    cell = cf_ws.cell(row=cf_rows["Depreciation"], column=col,
                      value=f"={pl_ref('Depreciation', col)}")
    cell.number_format = '#,##0_);(#,##0);"-"'; apply_link(cell)

    # Change in AR = −(AR_curr − AR_prev)
    ar_c = ref(BS, bs_rows["Accounts Receivable"], col)
    ar_p = ref(BS, bs_rows["Accounts Receivable"], col-1) if col > 2 else "0"
    cf_ws.cell(row=cf_rows["Change in AR"], column=col, value=f"=-({ar_c}-{ar_p})")
    apply_money(cf_ws.cell(row=cf_rows["Change in AR"], column=col))

    # Change in Inventory
    inv_c = ref(BS, bs_rows["Inventory"], col)
    inv_p = ref(BS, bs_rows["Inventory"], col-1) if col > 2 else "0"
    cf_ws.cell(row=cf_rows["Change in Inventory"], column=col, value=f"=-({inv_c}-{inv_p})")
    apply_money(cf_ws.cell(row=cf_rows["Change in Inventory"], column=col))

    # Change in AP
    ap_c = ref(BS, bs_rows["Accounts Payable"], col)
    ap_p = ref(BS, bs_rows["Accounts Payable"], col-1) if col > 2 else "0"
    cf_ws.cell(row=cf_rows["Change in AP"], column=col, value=f"={ap_c}-{ap_p}")
    apply_money(cf_ws.cell(row=cf_rows["Change in AP"], column=col))

    # Cash from Operations
    ni   = ref(CF, cf_rows["Net Income"], col)
    dep  = ref(CF, cf_rows["Depreciation"], col)
    dar  = ref(CF, cf_rows["Change in AR"], col)
    dinv = ref(CF, cf_rows["Change in Inventory"], col)
    dap  = ref(CF, cf_rows["Change in AP"], col)
    cf_ws.cell(row=cf_rows["Cash from Operations"], column=col,
               value=f"={ni}+{dep}+{dar}+{dinv}+{dap}")
    apply_money(cf_ws.cell(row=cf_rows["Cash from Operations"], column=col))

    # Capex = −(Revenue × capex_pct) → stored as negative
    cf_ws.cell(row=cf_rows["Capex"], column=col,
               value=f"=-({pl_ref('Revenue', col)}*{cp_ref('capex_pct')})")
    apply_money(cf_ws.cell(row=cf_rows["Capex"], column=col))

    # Cash from Investing = Capex
    cf_ws.cell(row=cf_rows["Cash from Investing"], column=col,
               value=f"={ref(CF, cf_rows['Capex'], col)}")
    apply_money(cf_ws.cell(row=cf_rows["Cash from Investing"], column=col))

    # Debt Repayment = −(Debt_curr − Debt_prev)
    debt_c = ref(BS, bs_rows["Debt"], col)
    debt_p = ref(BS, bs_rows["Debt"], col-1) if col > 2 else str(INITIAL_DEBT)
    cf_ws.cell(row=cf_rows["Debt Repayment"], column=col, value=f"=-({debt_c}-{debt_p})")
    apply_money(cf_ws.cell(row=cf_rows["Debt Repayment"], column=col))

    # Cash from Financing = Debt Repayment
    cf_ws.cell(row=cf_rows["Cash from Financing"], column=col,
               value=f"={ref(CF, cf_rows['Debt Repayment'], col)}")
    apply_money(cf_ws.cell(row=cf_rows["Cash from Financing"], column=col))

    # Net Change in Cash = Ops + Investing + Financing
    ops = ref(CF, cf_rows["Cash from Operations"], col)
    inv = ref(CF, cf_rows["Cash from Investing"], col)
    fin = ref(CF, cf_rows["Cash from Financing"], col)
    cf_ws.cell(row=cf_rows["Net Change in Cash"], column=col, value=f"={ops}+{inv}+{fin}")
    apply_money(cf_ws.cell(row=cf_rows["Net Change in Cash"], column=col))

# ─────────────────────────────────────────────
# 8. Valuation
# ─────────────────────────────────────────────
val = wb.create_sheet("Valuation")
apply_header(val.cell(row=1, column=1, value="DCF VALUATION (USD)"))
val.merge_cells(f"A1:{CL(YEARS+1)}1")

val["A3"] = "WACC";              val["B3"] = f"={cp_ref('wacc')}";             apply_percent(val["B3"]); apply_label(val["A3"], bold=True)
val["A4"] = "Perpetual Growth"; val["B4"] = f"={cp_ref('perpetual_growth')}"; apply_percent(val["B4"]); apply_label(val["A4"], bold=True)
WACC_REF = "'Valuation'!$B$3"
PERP_REF = "'Valuation'!$B$4"

apply_subheader(val.cell(row=6, column=1, value="Free Cash Flow to Firm (FCFF)"))
val.merge_cells(f"A6:{CL(YEARS+1)}6")
for c, y in enumerate(range(1, YEARS+1), start=2):
    apply_header(val.cell(row=7, column=c, value=f"Year {y}"))

for col in range(2, YEARS+2):
    ebit  = pl_ref("EBIT", col)
    depr  = pl_ref("Depreciation", col)
    capex = ref(CF, cf_rows["Capex"], col)
    dar   = ref(CF, cf_rows["Change in AR"], col)
    dinv  = ref(CF, cf_rows["Change in Inventory"], col)
    dap   = ref(CF, cf_rows["Change in AP"], col)
    tax   = cp_ref("tax_rate")
    val.cell(row=8, column=col,
             value=f"={ebit}*(1-{tax})+{depr}+{dar}+{dinv}+{dap}+{capex}")
    apply_money(val.cell(row=8, column=col))

last_fcff = f"'Valuation'!${CL(YEARS+1)}$8"
val["A10"] = "Terminal Value";  apply_label(val["A10"], bold=True)
val["B10"] = f"=IFERROR(({last_fcff}*(1+{PERP_REF}))/({WACC_REF}-{PERP_REF}),\"⚠ WACC ≤ g\")"
apply_money(val["B10"])

val["A11"] = "Enterprise Value"; apply_label(val["A11"], bold=True)
val["B11"] = (f"=IFERROR(NPV({WACC_REF},'Valuation'!${CL(2)}$8:${CL(YEARS+1)}$8)"
              f"+(B10/(1+{WACC_REF})^{YEARS}),0)")
apply_money(val["B11"])
val["B11"].font = _font(bold=True)

final_col = YEARS + 1
val["A12"] = "Less: Net Debt";   apply_label(val["A12"])
val["B12"] = f"={ref(BS, bs_rows['Debt'], final_col)}-{ref(BS, bs_rows['Cash'], final_col)}"
apply_money(val["B12"])

val["A13"] = "Equity Value";     apply_label(val["A13"], bold=True)
val["B13"] = "=B11-B12";         apply_money(val["B13"])
val["B13"].font = _font(bold=True, size=12)
val["B13"].border = border_top

# Implied multiples (small but high-value addition)
final_rev_val = pl_ref("Revenue", final_col)
final_ebitda_val = pl_ref("EBITDA", final_col)
val["A15"] = "Implied EV/Revenue (Yr 5)"; apply_label(val["A15"])
val["B15"] = f"=IFERROR(B11/{final_rev_val},0)"; val["B15"].number_format = "0.0x"; apply_formula(val["B15"])
val["A16"] = "Implied EV/EBITDA (Yr 5)";  apply_label(val["A16"])
val["B16"] = f"=IFERROR(B11/{final_ebitda_val},0)"; val["B16"].number_format = "0.0x"; apply_formula(val["B16"])

# ─────────────────────────────────────────────
# 9. Sensitivity (Excel Data Table)
# ─────────────────────────────────────────────
sens = wb.create_sheet("Sensitivity")

growth_rates = [0.02, 0.05, 0.08, 0.11, 0.14, 0.17, 0.20]
cogs_rates   = [0.35, 0.40, 0.45, 0.50, 0.55, 0.60, 0.65]
n_g = len(growth_rates)
n_c = len(cogs_rates)

# Title + instructions merged
apply_header(sens.cell(row=1, column=1,
    value=f"Net Income Sensitivity – Year {YEARS}"))
sens.merge_cells(f"A1:{CL(n_c+1)}1")

# Instructions row (row 2, spanning full width) — highlighted so users don't miss it
inst_cell = sens.cell(row=2, column=1,
    value=f"▶  Select B3:{CL(n_c+1)}{n_g+3}  →  Data  →  What-If Analysis  →  Data Table"
          f"   |   Row input: '{CP}'!$B${EFF_GROWTH_ROW}   |   Col input: '{CP}'!$B${EFF_COGS_ROW}")
inst_cell.font = _font(color="FFFFFF", bold=True, size=9)
inst_cell.fill = _fill("ED7D31")   # orange — hard to miss
inst_cell.alignment = Alignment(horizontal="left", vertical="center")
sens.merge_cells(f"A2:{CL(n_c+1)}2")
set_row_height(sens, 2, 22)

# Header row (row 3)
corner = sens.cell(row=3, column=1, value=f"={pl_ref('Net Income', YEARS+1)}")
apply_money(corner)
corner.fill = _fill("F2F2F2")

for j, cog in enumerate(cogs_rates, start=2):
    cell = sens.cell(row=3, column=j, value=cog)
    apply_percent(cell, is_input=True)
    cell.fill = _fill(LIGHT_BLUE := "D6E4F0")

for i, gr in enumerate(growth_rates, start=4):
    cell = sens.cell(row=i, column=1, value=gr)
    apply_percent(cell, is_input=True)
    cell.fill = _fill("D6E4F0")

# Body — placeholder dashes (Data Table fills these)
for i in range(n_g):
    for j in range(n_c):
        cell = sens.cell(row=4+i, column=2+j, value="—")
        cell.font      = _font(color="AAAAAA", size=9, italic=True)
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = '#,##0_);(#,##0);"-"'

# Heatmap conditional formatting (activates once Data Table runs)
heat_range = f"B4:{CL(n_c+1)}{n_g+3}"
sens.conditional_formatting.add(heat_range, ColorScaleRule(
    start_type="min",      start_color="F8696B",
    mid_type="percentile", mid_value=50, mid_color="FFEB84",
    end_type="max",        end_color="63BE7B"
))
sens.freeze_panes = "B4"

# ─────────────────────────────────────────────
# 10. Insights
# ─────────────────────────────────────────────
ins = wb.create_sheet("Insights")
apply_header(ins.cell(row=1, column=1, value="KEY PERFORMANCE INDICATORS"))
ins.merge_cells("A1:C1")
for c, h in enumerate(["Metric", f"Year {YEARS} Value", "Flag / Commentary"], start=1):
    apply_header(ins.cell(row=2, column=c, value=h))

final_rev    = pl_ref("Revenue",    YEARS+1)
yr1_rev      = pl_ref("Revenue",    2)
final_ni     = pl_ref("Net Income", YEARS+1)
final_ebitda = pl_ref("EBITDA",     YEARS+1)
final_cash   = ref(BS, bs_rows["Cash"], YEARS+1)
final_chg    = ref(CF, cf_rows["Net Change in Cash"], YEARS+1)

# Sources & Uses: Cash YN = Cash Y1 + sum(NChg Y2..YN)
# Cash rolls forward using CURRENT year NChg, so Y1's NChg is never included in the roll.
# Correct check: Ending Cash = Opening Cash + sum of NChg from Year 2 onward
cf_nchg_range = f"'Cash Flow'!${CL(3)}${cf_rows['Net Change in Cash']}:${CL(YEARS+1)}${cf_rows['Net Change in Cash']}"
src_uses_val  = f"IFERROR({ref(BS,bs_rows['Cash'],YEARS+1)}-({INITIAL_CASH}+SUMPRODUCT(({cf_nchg_range})*1)),0)"

# Break-even: first year where NI > 0
ni_range = f"'P&L'!$B${pl_rows['Net Income']}:${CL(YEARS+1)}${pl_rows['Net Income']}"
breakeven = f"IFERROR(MATCH(TRUE,INDEX({ni_range}>0,),0),\"Not in forecast\")"

kpis = [
    ("Revenue CAGR  (Yr 1 → 5)",
     f"=IFERROR(({final_rev}/{yr1_rev})^(1/{YEARS-1})-1,0)",
     "percent",
     f'=IF(\'Insights\'!B3>0.15,"🚀 Strong","=IF"&IF(\'Insights\'!B3>0.08,"✅ Solid","⚠ Below 8%"))'),

    ("EBITDA Margin",
     f"=IFERROR({final_ebitda}/{final_rev},0)",
     "percent",
     f'=IF(\'Insights\'!B4>=0.2,"✅ Healthy  (≥20%)","⚠ Below 20% — watch costs")'),

    ("Net Income Margin",
     f"=IFERROR({final_ni}/{final_rev},0)",
     "percent",
     f'=IF(\'Insights\'!B5>=0.1,"✅ Solid  (≥10%)","⚠ Below 10%")'),

    ("Gross Margin",
     f"=IFERROR(1-{EFF_COGS_REF},0)",
     "percent",
     f'=IF(1-{EFF_COGS_REF}>=0.5,"✅ Strong  (≥50%)","⚠ Below 50%")'),

    ("Profitable  (Year 5)?",
     f'=IF({final_ni}>0,"✅  Yes","❌  No")',
     "text", ""),

    ("Ending Cash",
     f"={final_cash}",
     "money",
     f'=IF({final_cash}>0,"✅ Positive","❌ NEGATIVE — model may be breaking")'),

    ("Cash Runway  (years)",
     f'=IFERROR(IF({final_chg}<0,{final_cash}/ABS({final_chg}),"∞  (positive CF)"),"N/A")',
     "text",
     "Years of cash at Year 5 burn rate"),

    ("Break-Even Year",
     f"={breakeven}",
     "text",
     "First year Net Income > 0"),

    ("Active Scenario",
     f"='{CP}'!{CP_SEL_ADDR}",
     "text",
     f'="{",".join(scen_names)} available"'),

    ("Active Revenue Growth",
     f"={EFF_GROWTH_REF}",
     "percent",
     f'=IF({EFF_GROWTH_REF}>0.18,"⚠ Very aggressive","✅ Plausible")'),

    ("Sources & Uses Check",
     f"={src_uses_val}",
     "money",
     f'=IF(ABS({src_uses_val})<1,"✅ Cash reconciles","❌ Cash reconciliation error")'),

    ("Balance Check",
     f"=SUM('Balance Sheet'!{BC_RANGE})",
     "money",
     f'=IF(ABS(SUM(\'Balance Sheet\'!{BC_RANGE}))<1,"✅ Balanced","❌ Out of balance — fix Balance Sheet")'),
]

for i, (label, formula, fmt, comment) in enumerate(kpis, start=3):
    ins.cell(row=i, column=1, value=label); apply_label(ins.cell(row=i, column=1), bold=True)
    ins.cell(row=i, column=2, value=formula)
    if fmt == "percent": apply_percent(ins.cell(row=i, column=2))
    elif fmt == "money": apply_money(ins.cell(row=i, column=2))
    else:                apply_formula(ins.cell(row=i, column=2))
    if comment:
        ins.cell(row=i, column=3, value=comment); apply_formula(ins.cell(row=i, column=3))

ins.column_dimensions["A"].width = 28
ins.column_dimensions["B"].width = 22
ins.column_dimensions["C"].width = 44

# ─────────────────────────────────────────────
# 11. Dashboard
# ─────────────────────────────────────────────
dash = wb.create_sheet("Dashboard")
dash["A1"] = "MODEL DASHBOARD"
dash["A1"].font = _font(bold=True, size=16, color=DARK_BLUE)
dash.merge_cells("A1:K1")
set_row_height(dash, 1, 30)

cats = Reference(pl, min_col=2, max_col=YEARS+1, min_row=1, max_row=1)

def make_bar(title, data_ref, color, anchor):
    chart = BarChart()
    chart.type = "col"; chart.grouping = "clustered"
    chart.title = title; chart.style = 10
    chart.y_axis.title = "USD"; chart.x_axis.title = "Year"
    chart.width = 14; chart.height = 10
    chart.add_data(data_ref)
    chart.series[0].graphicalProperties.solidFill = color
    chart.set_categories(cats)
    dash.add_chart(chart, anchor)

make_bar("Revenue  (USD)",
    Reference(pl, min_col=2, max_col=YEARS+1,
              min_row=pl_rows["Revenue"], max_row=pl_rows["Revenue"]),
    "2E75B6", "A3")

make_bar("Net Income  (USD)",
    Reference(pl, min_col=2, max_col=YEARS+1,
              min_row=pl_rows["Net Income"], max_row=pl_rows["Net Income"]),
    "70AD47", "H3")

# EBITDA Margin line — helper data written to a hidden row
MARGIN_ROW = 35
for col in range(2, YEARS+2):
    dash.cell(row=MARGIN_ROW, column=col,
              value=f"=IFERROR({pl_ref('EBITDA',col)}/{pl_ref('Revenue',col)},0)")
    apply_percent(dash.cell(row=MARGIN_ROW, column=col))
set_row_height(dash, MARGIN_ROW, 0.1)   # effectively hide the helper row

line = LineChart()
line.title = "EBITDA Margin %"; line.style = 10
line.y_axis.title = "Margin %"; line.y_axis.numFmt = "0%"
line.width = 14; line.height = 10
margin_data = Reference(dash, min_col=2, max_col=YEARS+1,
                        min_row=MARGIN_ROW, max_row=MARGIN_ROW)
line.add_data(margin_data)
line.series[0].title = SeriesLabel(v="EBITDA Margin")
line.series[0].graphicalProperties.line.solidFill = "ED7D31"
line.series[0].graphicalProperties.line.width = 25000
line.set_categories(cats)
dash.add_chart(line, "A20")

dash.sheet_view.showGridLines = False

# ─────────────────────────────────────────────
# 12. Model health — back-fill into Control Panel
# ─────────────────────────────────────────────
cp.cell(row=HEALTH_ROW, column=1, value="Model Health")
apply_label(cp.cell(row=HEALTH_ROW, column=1), bold=True)
cp.cell(row=HEALTH_ROW, column=2,
        value=f"=IF(ABS(SUM('Balance Sheet'!{BC_RANGE}))<1,\"✅  HEALTHY\",\"❌  BALANCE ERROR\")")
cp.cell(row=HEALTH_ROW, column=2).font = _font(bold=True, size=11)

# ─────────────────────────────────────────────
# 13. Start Here
# ─────────────────────────────────────────────
start = wb.create_sheet("Start Here")
start["A1"] = "📘  FINANCIAL MODEL"
start["A1"].font = _font(bold=True, size=20, color=DARK_BLUE)
start.merge_cells("A1:D1")
set_row_height(start, 1, 36)

# Sub-title
start["A2"] = "3-Statement model  ·  DCF Valuation  ·  Scenario Engine  ·  Sensitivity Analysis"
start["A2"].font = _font(size=10, color="595959", italic=True)
start.merge_cells("A2:D2")

steps = [
    ("Control Panel",  "Set assumptions (yellow cells) and pick a scenario from the dropdown. All sheets update."),
    ("P&L",            "Revenue → Gross Profit → EBITDA → Net Income across all forecast years."),
    ("Balance Sheet",  "Assets = Liabilities + Equity every year. Green balance check = model is correct."),
    ("Cash Flow",      "Indirect method: Net Income reconciled to ending cash. Ties back to Balance Sheet."),
    ("Valuation",      "Unlevered FCFF DCF → Terminal Value → Enterprise Value → Equity Value. Multiples shown."),
    ("Sensitivity",    "Data Table: vary Revenue Growth × COGS % to stress-test Year 5 Net Income."),
    ("Insights",       "KPIs — margins, CAGR, break-even year, cash runway, balance & reconciliation checks."),
    ("Dashboard",      "Charts: Revenue, Net Income, EBITDA Margin across all forecast years."),
]
start["A4"] = "Sheet"; start["C4"] = "What it does"
apply_header(start["A4"]); apply_header(start["C4"])
start.merge_cells("C4:D4")

for i, (tab, desc) in enumerate(steps, start=5):
    start.cell(row=i, column=1, value=f"  {i-4}.  {tab}")
    start.cell(row=i, column=1).font = _font(bold=True, color=DARK_BLUE)
    start.cell(row=i, column=3, value=desc)
    start.cell(row=i, column=3).font = _font()
    start.merge_cells(f"C{i}:D{i}")

start["A14"] = "Color coding"
start["A14"].font = _font(bold=True)
legend = [
    ("Blue text on yellow", "Hardcoded input — safe to change"),
    ("Black text",          "Formula — do not edit"),
    ("Green text",          "Cross-sheet link — do not edit"),
]
for r, (label, note) in enumerate(legend, start=15):
    start.cell(row=r, column=1, value=f"  {label}"); start.cell(row=r, column=1).font = _font()
    start.cell(row=r, column=3, value=note);         start.cell(row=r, column=3).font = _font()
    start.merge_cells(f"C{r}:D{r}")

start["A19"] = "Sensitivity note"
start["A19"].font = _font(bold=True)
start["C19"] = (f"After opening, go to Sensitivity sheet, select B3:{CL(n_c+1)}{n_g+3}, "
                f"Data → What-If Analysis → Data Table. "
                f"Row input = Control Panel B{EFF_GROWTH_ROW}, Col input = Control Panel B{EFF_COGS_ROW}.")
start["C19"].font = _font(italic=True, color="595959")
start.merge_cells("C19:D19")

# Metadata
for r, (k, v) in enumerate([("Version", "4.0"), ("Author", "Abdallah"), ("Date", "2026")], start=21):
    start.cell(row=r, column=1, value=k).font  = _font(bold=True, size=9, color="595959")
    start.cell(row=r, column=3, value=v).font  = _font(size=9, color="595959")

start.column_dimensions["A"].width = 24
start.column_dimensions["B"].width = 4
start.column_dimensions["C"].width = 56
start.sheet_view.showGridLines = False

# ─────────────────────────────────────────────
# 14. Global sheet styling
# ─────────────────────────────────────────────
data_sheets = [pl, bs_ws, cf_ws, val, sens, ins]
for ws in data_sheets:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B2"
    set_col_widths(ws, label_col_width=32, data_col_width=14)
    ws.protection.sheet = True
    set_row_height(ws, 1, 22)

# Sensitivity has different freeze
sens.freeze_panes = "B4"

cp.sheet_view.showGridLines = False
cp.freeze_panes = "A3"
cp.protection.sheet = True

# Unlock all inputs in Control Panel
for r in range(CP_ASSUMP_START, CP_ASSUMP_START + len(assumption_order)):
    cp.cell(row=r, column=2).protection = Protection(locked=False)
for r in range(SCEN_DATA_START, SCEN_DATA_START + len(SCENARIOS)):
    for c in range(1, 4):
        cp.cell(row=r, column=c).protection = Protection(locked=False)
cp.cell(row=SEL_ROW, column=2).protection = Protection(locked=False)

# ─────────────────────────────────────────────
# 15. Sheet order & cleanup
# ─────────────────────────────────────────────
order = ["Start Here", "Control Panel", "P&L", "Balance Sheet",
         "Cash Flow", "Valuation", "Sensitivity", "Insights", "Dashboard"]
wb._sheets = [wb[t] for t in order if t in wb.sheetnames]
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# ─────────────────────────────────────────────
# 16. Save
# ─────────────────────────────────────────────
out_dir  = Path(__file__).parent / "output"
out_dir.mkdir(exist_ok=True)
out_path = out_dir / f"Financial_Model.xlsx"
wb.save(out_path)

print(f"✅  Saved → {out_path}")
print(f"    Balance sheet will be GREEN on open.")
print(f"    Sensitivity → select B3:{CL(n_c+1)}{n_g+3}")
print(f"    → Data → What-If Analysis → Data Table")
print(f"    → Row input:  '{CP}'!$B${EFF_GROWTH_ROW}")
print(f"    → Col input:  '{CP}'!$B${EFF_COGS_ROW}")